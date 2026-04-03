import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import io
import json
import re
import requests
from datetime import datetime, date
import traceback

# ─── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Rent Roll AI Agent",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main { padding-top: 0.5rem; }
    .log-box {
        background: #0f1117;
        border: 1px solid #2d3748;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        font-family: 'Courier New', monospace;
        font-size: 0.78rem;
        max-height: 400px;
        overflow-y: auto;
        color: #e2e8f0;
        line-height: 1.6;
    }
    .log-step  { color: #63b3ed; }
    .log-think { color: #f6ad55; }
    .log-action{ color: #68d391; }
    .log-warn  { color: #fc8181; }
    .log-ok    { color: #9ae6b4; }
</style>
""", unsafe_allow_html=True)

TEMPLATE_PATH = "Rent_Roll_template.xlsx"

# ─── Rent charge-code patterns ───────────────────────────────────────────────
RENT_INCLUDE = [
    # ── Exact abbreviations ──────────────────────────────────────────────────
    r"^rent$",        # RENT
    r"^rnt$",         # RNT
    r"^rnta$",        # RNTA
    r"^rntb$",        # RNTB  (unit-type variant)
    r"^rntm$",        # RNTM  (market rent variant)
    r"^base$",        # BASE
    r"^brt$",         # BRT   (base rent)
    r"^br$",          # BR    (base rent short)
    r"^contract$",    # CONTRACT
    r"^crt$",         # CRT   (contract rent)
    r"^net$",         # NET
    r"^nr$",          # NR    (net rent)
    r"^hap$",         # HAP   (Housing Assistance Payment)
    r"^hapr$",        # HAPR  (HAP rent)
    r"^sec8$",        # SEC8
    r"^s8$",          # S8
    r"^s8r$",         # S8R
    r"^hud$",         # HUD
    r"^hudr$",        # HUDR
    r"^hudrnt$",      # HUDRNT
    r"^subsidy$",     # SUBSIDY
    r"^sub$",         # SUB
    r"^subs$",        # SUBS
    r"^ttp$",         # TTP   (Total Tenant Payment)
    r"^tenant$",      # TENANT
    r"^tr$",          # TR    (tenant rent)
    r"^tp$",          # TP    (tenant portion)
    r"^lihtc$",       # LIHTC (Low Income Housing Tax Credit)
    r"^tc$",          # TC    (tax credit)
    r"^credit$",      # CREDIT
    r"^bmr$",         # BMR   (Below Market Rate)
    r"^aff$",         # AFF   (Affordable)
    r"^usda$",        # USDA
    r"^rd$",          # RD    (Rural Development)
    r"^stl$",         # STL   (subsidized tenant lease)
    r"^rentmkt$",     # RENTMKT (Yardi market rent code)
    r"^mkt$",         # MKT
    r"^gr$",          # GR    (gross rent)
    r"^grs$",         # GRS
    r"^lr$",          # LR    (lease rent)
    r"^lrnt$",        # LRNT
    r"^cr$",          # CR    (contract rent)
    # ── Prefix patterns — catches "RENT-Rent", "HUDR-HUD Rent", etc. ────────
    r"^rent[-_]",     # RENT- prefix
    r"^rent:",         # RENT: prefix
    r"^hud[-_]",      # HUD-  prefix
    r"^hap[-_]",      # HAP-  prefix
    r"^s8[-_]",       # S8-   prefix
    r"^sec[-_]?8",    # SEC8 / SEC-8
    # ── Contains patterns — catches "Rent HUD", "County Rent", etc. ─────────
    r"\brent\b",        # word "rent" anywhere
    r"\blease\s*rent\b",
    r"\bbase\s*rent\b",
    r"\bcontract\s*rent\b",
    r"\bmarket\s*rent\b",
    r"\bnet\s*rent\b",
    r"\bgross\s*rent\b",
    r"\btenant\s*rent\b",
    r"\btenant\s*portion\b",
    r"\bhud\b",         # word "hud" anywhere  (catches "Rent HUD", "HUD Rent")
    r"\bhap\b",         # word "hap" anywhere
    r"\bsection\s*8\b",
    r"\bsubsidy\b",
    r"\bsubsidized\b",
    r"\bhousing\s*assist",
    r"\blihtc\b",
    r"\btax\s*credit\b",
    r"\baffordable\b",
    r"\busda\b",
    r"\brural\s*dev",
]
RENT_EXCLUDE = [
    r"pet", r"park", r"garag", r"storag", r"trash", r"water", r"sewer",
    r"electric", r"gas", r"internet", r"cable", r"utility", r"admin",
    r"pest", r"valet", r"amenity", r"concession", r"late\s*fee", r"nsf",
    r"real\s*estate\s*tax", r"pack", r"locker", r"waste",
    r"building\s*facilit", r"reimb", r"comfee", r"feeamty",
    r"ltor", r"ltol", r"loss\s*to\s*old", r"^real$",
]

def is_rent_code(code):
    if not code: return False
    cl = str(code).strip().lower()
    for ex in RENT_EXCLUDE:
        if re.search(ex, cl): return False
    for pat in RENT_INCLUDE:
        if re.search(pat, cl, re.IGNORECASE): return True
    return False

SKIP_KW  = ["pending renewal","future resident","applicant","future","prospect",
             "notice to vacate","previous","former"]
VACAT_STATUS_KW = ["vacant","model","down unit","offline","-- vacant --","excluded"]

def should_skip(status):
    sl = str(status or "").lower()
    return any(kw in sl for kw in SKIP_KW)

def is_vacant(status, name):
    sl = str(status or "").lower()
    nl = str(name or "").lower().strip()
    for kw in VACAT_STATUS_KW:
        if kw in sl: return True
    # Name-based placeholder detection
    if nl in ("-- vacant --", "vacant", "-"):
        return True
    # "Model R00000008" — starts with "model " (space after)
    if re.match(r"^(model|vacant)\s", nl):
        return True
    # "MODEL, MODEL" or "Model, Model" — all tokens are the same placeholder word
    name_words = re.split(r"[\s,]+", nl)
    if name_words and all(w in ("model","vacant","down") for w in name_words if w):
        return True
    return False

def fmt_date(val):
    if val is None: return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%m-%d-%Y")
    s = str(val).strip()
    for fmt in ["%m/%d/%Y","%Y-%m-%d","%m-%d-%Y","%d/%m/%Y","%m/%d/%y"]:
        try: return datetime.strptime(s, fmt).strftime("%m-%d-%Y")
        except: pass
    try:
        d = pd.to_datetime(s, errors="coerce")
        if not pd.isna(d): return d.strftime("%m-%d-%Y")
    except: pass
    return ""

def to_num(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    try: return float(str(val).replace(",","").replace("$","").strip())
    except: return None

def clean_type(val):
    if not val: return ""
    return re.sub(r"^unit\s*type\s*[:\-]\s*","",str(val),flags=re.IGNORECASE).strip()

def gcol(row, idx, default=None):
    if idx is None or idx >= len(row): return default
    return row[idx]

# ─── Claude API ───────────────────────────────────────────────────────────────
def call_claude(messages, system, max_tokens=2000):
    api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"Content-Type":"application/json","x-api-key": api_key,
                 "anthropic-version":"2023-06-01"},
        json={"model":"claude-sonnet-4-20250514","max_tokens":max_tokens,
              "system":system,"messages":messages},
    )
    data = resp.json()
    if "content" not in data: raise ValueError(f"API error: {data}")
    return "".join(b.get("text","") for b in data["content"] if b.get("type")=="text")

# ─── Parsers ─────────────────────────────────────────────────────────────────

def _finalize(current, units):
    if not current: return
    if current.get("_vacant"):
        current["effective_rent"] = None
    elif current.get("_charge_sum",0) > 0 and not current.get("effective_rent"):
        current["effective_rent"] = current["_charge_sum"]
    current.pop("_charge_sum", None)
    current.pop("_vacant", None)
    units.append(current)

def find_header(rows, keywords, max_search=20):
    for i, row in enumerate(rows[:max_search]):
        rs = " ".join(str(v or "").lower() for v in row)
        if all(kw in rs for kw in keywords):
            return i
    return None

def hdr_map(headers, *kws):
    for kw in kws:
        for i, h in enumerate(headers):
            if kw in str(h or "").lower().replace("\n"," "):
                return i
    return None


class YardiParser:
    """Yardi Voyager / Breeze: single header row, one data row per unit."""
    name = "Yardi"

    def can_handle(self, rows, fname):
        for row in rows[:12]:
            rs = " ".join(str(v or "").lower() for v in row)
            if ("resh id" in rs or "unit/lease status" in rs or
                ("unit" in rs and "floorplan" in rs and "sqft" in rs and "lease rent" in rs)):
                return True
        return False

    def extract(self, rows):
        hi = find_header(rows, ["unit","status"], max_search=15)
        if hi is None: return []
        h = [str(v or "").lower().replace("\n"," ").strip() for v in rows[hi]]
        c = dict(
            unit   = hdr_map(h,"unit","bldg-unit"),
            fp     = hdr_map(h,"floorplan","unit type","floor plan"),
            sqft   = hdr_map(h,"sqft","sq ft","size"),
            status = hdr_map(h,"status","designation"),
            name   = hdr_map(h,"name"),
            movin  = hdr_map(h,"move-in","move in"),
            lstart = hdr_map(h,"lease start","lease\nstart"),
            lend   = hdr_map(h,"lease end","lease\nend"),
            mkt    = hdr_map(h,"market rent","market\nrent"),
            eff    = hdr_map(h,"lease rent","effective rent","actual rent"),
        )
        units = []
        for row in rows[hi+1:]:
            # Detect mid-file section dividers / repeated header rows in OneSite
            r1_val = row[1] if len(row) > 1 else None
            r1_str = str(r1_val or "").strip()
            if r1_str == "details" or r1_str == "\nUnit" or (
                r1_val is None and len(row) > 36 and row[36] and 
                "trans" in str(row[36]).lower()
            ):
                # Mid-file section divider - skip but keep accumulating for current unit
                continue

            uv = gcol(row, c["unit"])
            if not uv: continue
            uvs = str(uv).strip()
            if not uvs or uvs.lower() in ["unit","none","nan"]: continue
            if any(kw in uvs.lower() for kw in ["total","summary","subtotal","grand","future residents"]): break
            status = str(gcol(row, c["status"]) or "")
            name   = str(gcol(row, c["name"])   or "")
            if should_skip(status): continue
            vacat  = is_vacant(status, name)
            eff    = to_num(gcol(row, c["eff"]))
            if vacat: eff = None
            units.append({
                "unit_no":       uvs,
                "unit_type":     clean_type(gcol(row, c["fp"])),
                "sqft":          to_num(gcol(row, c["sqft"])),
                "status":        status,
                "resident_name": name if not vacat else "",
                "move_in":       fmt_date(gcol(row, c["movin"])),
                "lease_start":   fmt_date(gcol(row, c["lstart"])),
                "lease_end":     fmt_date(gcol(row, c["lend"])),
                "market_rent":   to_num(gcol(row, c["mkt"])),
                "effective_rent": eff,
                "building": None,
            })
        return units


class OneSiteParser:
    """RealPage OneSite: wide columns (58), unit row + charge rows, \n in headers."""
    name = "OneSite/RealPage"

    def can_handle(self, rows, fname):
        for row in rows[:5]:
            rs = " ".join(str(v or "").lower() for v in row)
            if "onesite" in rs or "onsite" in rs: return True
        return False

    def _find_header_row(self, rows):
        for i, row in enumerate(rows[:15]):
            for v in row:
                if v and "\nUnit" in str(v) or (v and "unit" in str(v).lower() and "\n" in str(v)):
                    return i
            # Also check row with Trans Code
            rs = " ".join(str(v or "") for v in row)
            if "Trans" in rs and "Code" in rs and "Floorplan" in rs:
                return i
        return None

    def extract(self, rows):
        hi = self._find_header_row(rows)
        if hi is None: return []
        h = rows[hi]
        c = dict(
            unit   = hdr_map(h,"\nunit","unit\n"),
            fp     = hdr_map(h,"floorplan","floor plan"),
            sqft   = hdr_map(h,"sqft","sq ft"),
            status = hdr_map(h,"unit/lease\nstatus","status"),
            name   = hdr_map(h,"name\n","\nname"),
            movin  = hdr_map(h,"move-in\nmove-out","move-in"),
            lstart = hdr_map(h,"lease\nstart"),
            lend   = hdr_map(h,"lease\nend"),
            mkt    = hdr_map(h,"market\n+ addl.","market rent","market\nrent"),
            tcode  = hdr_map(h,"trans\ncode","code"),
            lrent  = hdr_map(h,"lease\nrent","lease rent"),
        )
        # Fallback column positions for known OneSite layout
        if c["unit"] is None:   c["unit"]   = 1
        if c["fp"] is None:     c["fp"]     = 3
        if c["sqft"] is None:   c["sqft"]   = 14
        if c["status"] is None: c["status"] = 18
        if c["name"] is None:   c["name"]   = 20
        if c["movin"] is None:  c["movin"]  = 24
        if c["lstart"] is None: c["lstart"] = 28
        if c["lend"] is None:   c["lend"]   = 30
        if c["mkt"] is None:    c["mkt"]    = 33
        if c["tcode"] is None:  c["tcode"]  = 36
        if c["lrent"] is None:  c["lrent"]  = 43

        # OneSite has a known column-offset artifact after unmerging:
        # 'Lease Rent' header may land at idx N but actual values sit at idx N+1.
        # Probe first charge row to auto-detect the real data column.
        for probe_row in rows[hi+1:hi+50]:
            tcode_val = gcol(probe_row, c["tcode"])
            if tcode_val and is_rent_code(str(tcode_val)):
                for offset in range(c["tcode"]+1, min(c["tcode"]+12, len(probe_row))):
                    v = probe_row[offset]
                    if v is not None and isinstance(v, (int, float)) and float(v) > 0:
                        c["lrent"] = offset
                        break
                break

        units = []
        current = None
        for row in rows[hi+1:]:
            # Skip mid-file section-divider rows (OneSite re-prints "details" + headers mid-file)
            r1_raw = row[1] if len(row) > 1 else None
            r1_str = str(r1_raw or "").strip()
            if r1_str in ("details",) or r1_str == "\nUnit":
                continue   # keep current unit alive, charges continue after header

            uv = gcol(row, c["unit"])
            if uv:
                uvs = str(uv).strip()
                if not uvs or uvs.lower() in ["unit","none","nan"]: continue
                if any(kw in uvs.lower() for kw in ["total","summary"]): break
                _finalize(current, units)
                name   = str(gcol(row, c["name"]) or "")
                status = str(gcol(row, c["status"]) or "")
                if should_skip(status): current=None; continue
                vacat  = is_vacant(status, name)
                movin_raw = re.split(r"_x000D_|\\r", str(gcol(row, c["movin"]) or ""))[0].strip()
                sqft_raw  = str(gcol(row, c["sqft"]) or "").replace(",","")
                cc = str(gcol(row, c["tcode"]) or "")
                ca = to_num(gcol(row, c["lrent"]))
                current = {
                    "unit_no":       uvs,
                    "unit_type":     clean_type(gcol(row, c["fp"])),
                    "sqft":          to_num(sqft_raw) if sqft_raw else None,
                    "status":        status,
                    "resident_name": name if not vacat else "",
                    "move_in":       fmt_date(movin_raw),
                    "lease_start":   fmt_date(gcol(row, c["lstart"])),
                    "lease_end":     fmt_date(gcol(row, c["lend"])),
                    "market_rent":   to_num(gcol(row, c["mkt"])),
                    "effective_rent": None,
                    "_charge_sum":   0,
                    "_vacant":       vacat,
                    "building": None,
                }
                if cc and ca and is_rent_code(cc): current["_charge_sum"] += ca or 0
            else:
                if current is None: continue
                # Skip future-tenant sub-rows embedded in current unit block
                row_str = " ".join(str(v or "").lower() for v in row if v)
                if any(kw in row_str for kw in ["applicant","pending renewal","pending renew"]):
                    current["_skip_future"] = True
                    continue
                if current.get("_skip_future"):
                    continue
                cc = str(gcol(row, c["tcode"]) or "")
                ca = to_num(gcol(row, c["lrent"]))
                if cc and ca and is_rent_code(cc): current["_charge_sum"] += ca or 0
        _finalize(current, units)
        return units


class MRIParser:
    """MRI Living: 13-14 col, 2-row headers.
    Variant A (14 col): multi-row charges, col6=ChargeCode, col7=Amount (Stone_Loch, Canal)
    Variant B (13 col): single effective rent, col6=ActualRent (Retreat_at_the_Park)
    """
    name = "MRI Living"

    def can_handle(self, rows, fname):
        # MRI has a 2-row split header: row N has "Unit|Unit Type|Unit|Resident|Name|Market|Charge|Amount"
        # and row N+1 has "Sq Ft" in position 2 completing the split.
        # Also: title row contains "Rent Roll with Lease Charges" or "As Of ="
        has_title = any("rent roll with lease charges" in " ".join(str(v or "").lower() for v in r) or
                        "as of =" in " ".join(str(v or "").lower() for v in r)
                        for r in rows[:6])
        if not has_title:
            return False
        for i, row in enumerate(rows[3:10], 3):
            rs = " ".join(str(v or "").lower() for v in row)
            if ("unit" in rs and ("charge" in rs or "actual" in rs) and "market" in rs
                    and "bldg-unit" not in rs and "bldg" not in rs):
                # Check next row for "sq ft"
                if i+1 < len(rows):
                    r2 = " ".join(str(v or "").lower() for v in rows[i+1])
                    if "sq ft" in r2:
                        return True
        return False

    def _detect_variant(self, rows):
        """Returns 'charges' or 'actual' based on header row."""
        for row in rows[4:8]:
            rs = " ".join(str(v or "").lower() for v in row)
            if "charge" in rs and "amount" in rs:
                return "charges"
            if "actual" in rs:
                return "actual"
        return "charges"

    def _find_data_start(self, rows):
        for i, row in enumerate(rows[:20]):
            r0 = str(row[0] or "").strip()
            if r0 and row[4] and not any(kw in r0.lower() for kw in
               ["rent roll","as of","month","unit","current","notice","vacant","total","summary"]):
                if re.match(r"[\w\d\-]+", r0):
                    return i
        return 7

    def extract(self, rows):
        variant    = self._detect_variant(rows)
        data_start = self._find_data_start(rows)

        units = []
        current = None

        for row in rows[data_start:]:
            r0 = str(row[0] or "").strip()
            r4 = str(row[4] or "").strip() if len(row) > 4 else ""
            if any(kw in r0.lower() for kw in ["total","summary","grand","future residents","future resident"]): break
            # Section headers
            if r0 and not r4 and all(row[j] is None for j in range(1, min(5, len(row)))):
                continue

            is_new = bool(r0 and r4)
            if is_new:
                _finalize(current, units)
                name  = r4
                vacat = is_vacant("", name)
                mkt   = to_num(row[5]) if len(row) > 5 else None

                if variant == "actual":
                    # col 6 = actual rent amount directly
                    eff = to_num(row[6]) if len(row) > 6 else None
                    if vacat: eff = None
                    units.append({
                        "unit_no":       r0,
                        "unit_type":     clean_type(row[1] if len(row) > 1 else None),
                        "sqft":          to_num(row[2]) if len(row) > 2 else None,
                        "status":        "",
                        "resident_name": name if not vacat else "",
                        "move_in":       fmt_date(row[9] if len(row) > 9 else None),
                        "lease_start":   "",
                        "lease_end":     fmt_date(row[10] if len(row) > 10 else None),
                        "market_rent":   mkt,
                        "effective_rent": eff,
                        "building": None,
                    })
                    current = None
                else:
                    # Variant A: multi-row charges
                    cc = str(row[6] or "").strip() if len(row) > 6 else ""
                    ca = to_num(row[7]) if len(row) > 7 else None
                    current = {
                        "unit_no":       r0,
                        "unit_type":     clean_type(row[1] if len(row) > 1 else None),
                        "sqft":          to_num(row[2]) if len(row) > 2 else None,
                        "status":        "",
                        "resident_name": name if not vacat else "",
                        "move_in":       fmt_date(row[10] if len(row) > 10 else None),
                        "lease_start":   "",
                        "lease_end":     fmt_date(row[11] if len(row) > 11 else None),
                        "market_rent":   mkt,
                        "effective_rent": None,
                        "_charge_sum":   0,
                        "_vacant":       vacat,
                        "building": None,
                    }
                    if cc and ca and is_rent_code(cc): current["_charge_sum"] += ca or 0
            else:
                if current is None or variant == "actual": continue
                cc = str(row[6] or "").strip() if len(row) > 6 else ""
                ca = to_num(row[7]) if len(row) > 7 else None
                if cc and ca and "total" not in cc.lower() and is_rent_code(cc):
                    current["_charge_sum"] += ca or 0

        if variant == "charges":
            _finalize(current, units)
        return units


class AppFolioParser:
    """AppFolio / ReNew style: unit type in section headers, multi-row charges."""
    name = "AppFolio"

    def can_handle(self, rows, fname):
        for row in rows[:12]:
            rs = " ".join(str(v or "").lower() for v in row)
            if ("bldg-unit" in rs or "unit details" in rs) and ("charge code" in rs or "scheduled" in rs):
                return True
        return False

    def extract(self, rows):
        hi = find_header(rows, ["bldg-unit","resident","lease"], max_search=15)
        if hi is None:
            hi = find_header(rows, ["bldg","resident","lease start"], max_search=15)
        if hi is None: return []

        h = [str(v or "").lower().replace("\n"," ").strip() for v in rows[hi]]
        c = dict(
            unit   = hdr_map(h,"bldg-unit","unit"),
            sqft   = hdr_map(h,"sqft","sq ft"),
            status = hdr_map(h,"unit status","status"),
            res    = hdr_map(h,"resident"),
            movin  = hdr_map(h,"move-in","move in"),
            lstart = hdr_map(h,"lease start"),
            lend   = hdr_map(h,"lease end"),
            mkt    = hdr_map(h,"budgeted rent","market rent","market"),
            cc     = hdr_map(h,"charge code"),
            sched  = hdr_map(h,"scheduled charges","scheduled"),
            ledger = hdr_map(h,"ledger"),
        )

        units = []
        current_type = None
        current = None

        for row in rows[hi+1:]:
            r0 = str(row[0] or "").strip()
            # Section header: single populated cell that is not a unit number
            if r0 and all(row[j] is None for j in range(1, min(6, len(row)))):
                if "future" in r0.lower():
                    break   # Stop at "Future Resident Details" / "Future Residents" section
                if not re.match(r"^\d", r0) and "total" not in r0.lower():
                    current_type = clean_type(r0)
                    continue

            # "Charge Total:" skip
            ledger_val = str(gcol(row, c["ledger"]) or "").lower()
            if "charge total" in ledger_val: continue

            uv = gcol(row, c["unit"])
            if not uv:
                # Charge-only row (no unit number) - accumulate into current unit
                if current is None: continue
                cc_v = str(gcol(row, c["cc"]) or "")
                ca_v = to_num(gcol(row, c["sched"]))
                if cc_v and ca_v and is_rent_code(cc_v): current["_charge_sum"] += ca_v or 0
                continue

            uvs = str(uv).strip()
            if not uvs or uvs.lower() in ["unit","bldg-unit","none","nan"]: continue
            if any(kw in uvs.lower() for kw in ["future residents","future resident",
                                                        "future resident details"]): break
            # Skip section subtotals like "Auburn Total:" - only stop on grand/property totals
            if re.match(r"^(grand total|property total|total units|total$)", uvs.lower()): break
            if uvs.lower() in ["total","summary","subtotal"]: break

            res_val    = gcol(row, c["res"])
            status_val = str(gcol(row, c["status"]) or "")
            is_unit_row = bool(status_val or res_val)

            if is_unit_row and re.match(r"[\w\d\-]", uvs):
                _finalize(current, units)
                if should_skip(status_val): current=None; continue
                name  = str(res_val or "")
                vacat = is_vacant(status_val, name)
                cc_v  = str(gcol(row, c["cc"])    or "")
                ca_v  = to_num(gcol(row, c["sched"]))
                current = {
                    "unit_no":       uvs,
                    "unit_type":     current_type or "",
                    "sqft":          to_num(gcol(row, c["sqft"])),
                    "status":        status_val,
                    "resident_name": name if not vacat else "",
                    "move_in":       fmt_date(gcol(row, c["movin"])),
                    "lease_start":   fmt_date(gcol(row, c["lstart"])),
                    "lease_end":     fmt_date(gcol(row, c["lend"])),
                    "market_rent":   to_num(gcol(row, c["mkt"])),
                    "effective_rent": None,
                    "_charge_sum":   0,
                    "_vacant":       vacat,
                    "building": None,
                }
                if cc_v and ca_v and is_rent_code(cc_v): current["_charge_sum"] += ca_v or 0
            else:
                # Row has a unit number but no status/res -> inline charge row
                if current is None: continue
                cc_v = str(gcol(row, c["cc"]) or "")
                ca_v = to_num(gcol(row, c["sched"]))
                if cc_v and ca_v and is_rent_code(cc_v): current["_charge_sum"] += ca_v or 0
        _finalize(current, units)
        return units


class RentManagerParser:
    """Rent Manager / WAT: cols Unit, Type, Market, Name+ID+Date, LeaseDates, ..., Desc, ChargeAmt."""
    name = "Rent Manager"

    def can_handle(self, rows, fname):
        for row in rows[:15]:
            rs = " ".join(str(v or "").lower() for v in row)
            if "lease dates" in rs and ("description" in rs or "charge amount" in rs):
                return True
        return False

    def extract(self, rows):
        hi = find_header(rows, ["lease dates","description"], max_search=20)
        if hi is None: return []

        units = []
        current = None

        for row in rows[hi+1:]:
            r0 = str(row[0] or "").strip()
            if any(kw in r0.lower() for kw in ["total","summary","grand","future residents","future resident"]): break
            is_unit = bool(r0 and re.match(r"^\d+", r0) and len(row) > 1 and row[1])
            if is_unit:
                _finalize(current, units)
                unit_no   = r0
                unit_type = str(row[1] or "").split(",")[0].strip()
                mkt       = to_num(row[2]) if len(row) > 2 else None
                name_raw  = str(row[3] or "")
                name_m    = re.match(r"^([^T\d]+?)(?:\s+T\d+|$)", name_raw)
                name      = name_m.group(1).strip() if name_m else name_raw.split("T00")[0].strip()
                dates_str = str(row[4] or "") if len(row) > 4 else ""
                dates     = re.findall(r"\d{2}/\d{2}/\d{4}", dates_str)
                vacat     = is_vacant("", name)
                desc      = str(row[8] or "") if len(row) > 8 else ""
                ca        = to_num(row[9]) if len(row) > 9 else None
                current = {
                    "unit_no":       unit_no,
                    "unit_type":     unit_type,
                    "sqft":          None,
                    "status":        "",
                    "resident_name": name if not vacat else "",
                    "move_in":       "",
                    "lease_start":   fmt_date(dates[0]) if len(dates) > 0 else "",
                    "lease_end":     fmt_date(dates[1]) if len(dates) > 1 else "",
                    "market_rent":   mkt,
                    "effective_rent": None,
                    "_charge_sum":   0,
                    "_vacant":       vacat,
                    "building": None,
                }
                if desc and ca and is_rent_code(desc): current["_charge_sum"] += ca or 0
            else:
                if current is None: continue
                desc = str(row[8] or "") if len(row) > 8 else ""
                ca   = to_num(row[9]) if len(row) > 9 else None
                if desc and ca and is_rent_code(desc): current["_charge_sum"] += ca or 0
        _finalize(current, units)
        return units


class ResProParser:
    """
    ResProp Management / Entrata-style format.
    Single sheet named 'Sheet'. Header at row 9.
    Cols: A=unit, C=type, E=sqft, F=name, K=status, M=market_rent,
          T=charge_desc, W=charge_amount, Z=move_in, AB=lease_start, AC=lease_end.
    Charge rows: only T and W populated (A-M all None).
    Vacant units identified by name='Vacant Unit' or empty status.
    Stop parsing at 'Total Charges' summary row.
    """
    name = "ResProp/Entrata"

    def can_handle(self, rows, fname):
        # Detect: single sheet named Sheet, header at row 9 has "Sq. Feet" and "Ledger"
        # and data rows have description in col 19 and amount in col 22
        for row in rows[8:12]:
            rs = " ".join(str(v or "").lower() for v in row)
            if "sq. feet" in rs and "ledger" in rs and "description" in rs:
                return True
        # Also detect by title rows: "ResProp" or "Rent Roll" with date format
        for row in rows[:7]:
            rs = " ".join(str(v or "") for v in row)
            if "ResProp" in rs or "Rent Roll" in rs:
                for r2 in rows[:7]:
                    for v in r2:
                        if v and "sq. feet" in str(v).lower():
                            return True
        return False

    def extract(self, rows):
        units   = []
        current = None

        def finalize():
            if current:
                if current.get("_vacant"):
                    current["effective_rent"] = None
                elif current.get("_charge_sum", 0) > 0:
                    current["effective_rent"] = current["_charge_sum"]
                current.pop("_charge_sum", None)
                current.pop("_vacant", None)
                units.append(current)

        for i, row in enumerate(rows):
            r0 = str(row[0] or "").strip()

            # Stop at summary section
            if r0 in ("Total Charges", "Description") or r0.startswith("Amenity Fees"):
                break

            # Unit row: col 0 has unit number, col 5 has resident name OR "Vacant Unit"
            is_unit = (
                row[0] is not None and
                (row[5] is not None or str(row[0]).strip().isdigit() or
                 re.match(r"^\d", str(row[0]).strip()))
                and row[2] is not None   # type must be present
            )

            if is_unit:
                finalize()
                name   = str(row[5] or "").strip()
                status = str(row[10] or "").strip()
                vacat  = is_vacant(status, name)

                if should_skip(status):
                    current = None
                    continue

                # Inline charge on unit row
                cc = str(row[19] or "").strip()
                ca = to_num(row[22])

                current = {
                    "unit_no":       r0,
                    "unit_type":     clean_type(str(row[2] or "")),
                    "sqft":          to_num(row[4]),
                    "status":        status,
                    "resident_name": name if not vacat else "",
                    "move_in":       fmt_date(row[26]),
                    "lease_start":   fmt_date(row[28]),
                    "lease_end":     fmt_date(row[29]),
                    "market_rent":   to_num(row[12]),
                    "effective_rent": None,
                    "_charge_sum":   0,
                    "_vacant":       vacat,
                    "building":      None,
                }
                if cc and ca and is_rent_code(cc):
                    current["_charge_sum"] += ca or 0

            elif row[0] is None and row[19]:
                # Charge or total row
                if current is None:
                    continue
                desc = str(row[19]).strip()
                if desc.lower() == "total":
                    continue   # separator row, skip
                ca = to_num(row[22])
                if ca and is_rent_code(desc):
                    current["_charge_sum"] += ca or 0

        finalize()
        return units


class AIFallbackParser:
    """Uses Claude to map any unknown format, then generic extraction."""
    name = "AI (Claude)"

    def detect_schema(self, rows, filename):
        preview = []
        seen = 0
        for i, r in enumerate(rows):
            if any(v is not None for v in r):
                preview.append(f"Row {i}: {[str(v)[:35] if v is not None else None for v in r[:28]]}")
                seen += 1
            if seen >= 45: break

        system = "You are an expert multi-family rent roll data analyst. Return ONLY valid JSON. No markdown."
        prompt = f"""Analyze raw rent roll from '{filename}'.

{chr(10).join(preview)}

Return JSON:
{{
  "format": "software name",
  "header_row": <0-based int>,
  "data_start_row": <0-based int>,
  "structure": "one_row_per_unit" | "multi_row_charges",
  "cols": {{
    "unit_no": <int|null>, "unit_type": <int|null>, "sqft": <int|null>,
    "status": <int|null>, "resident_name": <int|null>,
    "move_in": <int|null>, "lease_start": <int|null>, "lease_end": <int|null>,
    "market_rent": <int|null>, "effective_rent": <int|null>,
    "charge_code": <int|null>, "charge_amount": <int|null>
  }},
  "building_col": <int|null>,
  "unit_type_in_section_header": true|false,
  "notes": "brief"
}}"""
        try:
            raw = call_claude([{"role":"user","content":prompt}], system, 1500)
            raw = re.sub(r"```[a-z]*\n?","",raw).strip("` \n")
            return json.loads(raw)
        except Exception as e:
            return {
                "format":"unknown","header_row":0,"data_start_row":1,
                "structure":"multi_row_charges",
                "cols":{"unit_no":0,"unit_type":1,"sqft":2,"status":3,
                        "resident_name":4,"move_in":5,"lease_start":6,
                        "lease_end":7,"market_rent":8,"effective_rent":9,
                        "charge_code":None,"charge_amount":None},
                "building_col":None,"unit_type_in_section_header":False,"notes":str(e),
            }

    def extract(self, rows, schema):
        cols       = schema.get("cols",{})
        structure  = schema.get("structure","multi_row_charges")
        data_start = schema.get("data_start_row",1)
        cur_bldg   = None
        units      = []
        current    = None

        def g(row, key, default=None):
            idx = cols.get(key)
            if idx is None or idx >= len(row): return default
            return row[idx]

        for i, row in enumerate(rows[data_start:], data_start):
            rstr = " ".join(str(v or "").lower() for v in row)
            if re.search(r"\b(grand total|total units|summary)\b", rstr): break

            bc = schema.get("building_col")
            if bc is not None and bc < len(row) and row[bc]:
                if all(row[j] is None for j in range(len(row)) if j != bc):
                    cur_bldg = str(row[bc]).strip()
                    continue

            uv = g(row,"unit_no")
            if not uv: continue
            uvs = str(uv).strip()
            if not uvs or uvs.lower() in ["unit","none","nan"]: continue
            if any(kw in uvs.lower() for kw in ["total","summary","subtotal"]): continue

            status = str(g(row,"status") or "")
            name   = str(g(row,"resident_name") or "")

            if structure == "one_row_per_unit":
                if should_skip(status): continue
                vacat = is_vacant(status, name)
                eff   = to_num(g(row,"effective_rent"))
                if vacat: eff = None
                units.append({
                    "unit_no":       uvs,
                    "unit_type":     clean_type(g(row,"unit_type") or cur_bldg),
                    "sqft":          to_num(g(row,"sqft")),
                    "status":        status,
                    "resident_name": name if not vacat else "",
                    "move_in":       fmt_date(g(row,"move_in")),
                    "lease_start":   fmt_date(g(row,"lease_start")),
                    "lease_end":     fmt_date(g(row,"lease_end")),
                    "market_rent":   to_num(g(row,"market_rent")),
                    "effective_rent": eff,
                    "building": cur_bldg,
                })
            else:  # multi_row_charges
                is_new = bool(uv and (g(row,"resident_name") or g(row,"market_rent")))
                if is_new:
                    _finalize(current, units)
                    if should_skip(status): current=None; continue
                    vacat = is_vacant(status, name)
                    cc = str(g(row,"charge_code") or "")
                    ca = to_num(g(row,"charge_amount"))
                    current = {
                        "unit_no":       uvs,
                        "unit_type":     clean_type(g(row,"unit_type") or cur_bldg),
                        "sqft":          to_num(g(row,"sqft")),
                        "status":        status,
                        "resident_name": name if not vacat else "",
                        "move_in":       fmt_date(g(row,"move_in")),
                        "lease_start":   fmt_date(g(row,"lease_start")),
                        "lease_end":     fmt_date(g(row,"lease_end")),
                        "market_rent":   to_num(g(row,"market_rent")),
                        "effective_rent": to_num(g(row,"effective_rent")),
                        "_charge_sum":   0,
                        "_vacant":       vacat,
                        "building": cur_bldg,
                    }
                    if cc and ca and is_rent_code(cc): current["_charge_sum"] += ca or 0
                else:
                    if current is None: continue
                    cc = str(g(row,"charge_code") or "")
                    ca = to_num(g(row,"charge_amount"))
                    if cc and ca and is_rent_code(cc): current["_charge_sum"] += ca or 0

        if structure == "multi_row_charges":
            _finalize(current, units)
        return units


# ─── Agent ───────────────────────────────────────────────────────────────────
class RentRollAgent:

    PARSERS = [
        OneSiteParser(),
        RentManagerParser(),
        MRIParser(),
        AppFolioParser(),
        ResProParser(),
        YardiParser(),
    ]

    def __init__(self, log_fn=None):
        self.log = log_fn or (lambda l, m: None)

    def _find_raw_sheet(self, wb):
        for s in wb.sheetnames:
            if "raw" in s.lower(): return s
        for s in wb.sheetnames:
            if "rent roll" not in s.lower(): return s
        return wb.sheetnames[0]

    def _to_rows(self, ws, max_rows=3000):
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True):
            rows.append(list(row))
        return rows

    def _validate(self, units):
        if not units: return {"ok":False,"msg":"No units extracted"}
        eff = [u["effective_rent"] for u in units if u.get("effective_rent")]
        if not eff: return {"ok":False,"msg":"All units have zero effective rent"}
        avg = sum(eff)/len(eff)
        if avg < 50:  return {"ok":False,"msg":f"Avg eff rent ${avg:.0f} too low"}
        if avg > 100000: return {"ok":False,"msg":f"Avg eff rent ${avg:.0f} too high"}
        return {"ok":True,"msg":f"{len(units)} units · {len(eff)} occupied · avg ${avg:,.0f}"}

    def _write_template(self, units, rent_roll_date):
        """
        Write data directly into the template XML without openpyxl touching formulas.
        Only modifies columns C D H I J K N O Q (and G5 for the date).
        All formula columns, array formulas, and the file structure are preserved bit-for-bit.
        """
        import zipfile as _zf, re as _re, io as _io
        from datetime import date as _date

        def _serial(dt):
            if isinstance(dt, datetime): dt = dt.date()
            return (dt - _date(1899, 12, 30)).days

        def _esc(s):
            return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")

        # Column metadata — only these 9 data cols + G5 are ever touched
        COL_LTR   = {3:"C",4:"D",7:"G",8:"H",9:"I",10:"J",11:"K",14:"N",15:"O",17:"Q"}
        DATE_COLS = {7,9,10,11}
        STR_COLS  = {3,4,14}
        COL_STY   = {3:46,4:46,7:16,8:46,9:49,10:50,11:51,14:53,15:54,17:46}

        def _cell(col, row, value):
            ref   = f"{COL_LTR[col]}{row}"
            style = COL_STY.get(col, 46)
            if value is None or value == "":
                return f'<c r="{ref}" s="{style}"/>'
            if col in DATE_COLS:
                s = _serial(value) if isinstance(value, (datetime, _date)) else int(value)
                return f'<c r="{ref}" s="{style}"><v>{s}</v></c>'
            elif col in STR_COLS:
                return f'<c r="{ref}" s="{style}" t="inlineStr"><is><t>{_esc(str(value))}</t></is></c>'
            else:
                return f'<c r="{ref}" s="{style}"><v>{int(round(float(value)))}</v></c>'

        def _replace(xml, col, row, value):
            ref = f"{COL_LTR[col]}{row}"
            new = _cell(col, row, value)
            # Match self-closing OR element whose body does NOT start another <c> element
            p_sc = rf'<c r="{_re.escape(ref)}"(?:\s[^>]*)*/>' 
            p_ct = rf'<c r="{_re.escape(ref)}"(?:\s[^>]*)?>(?:(?!<c\s)[\s\S])*?</c>'
            return _re.sub(rf'(?:{p_ct}|{p_sc})', lambda m: new, xml, count=1)

        # Load template as raw zip bytes (preserves every formula, style, external-link-free)
        with _zf.ZipFile(TEMPLATE_PATH, 'r') as z:
            tpl = {name: z.read(name) for name in z.namelist()}

        xml = tpl['xl/worksheets/sheet1.xml'].decode('utf-8')

        # G5 — Rent Roll date
        try:
            xml = _replace(xml, 7, 5, datetime.strptime(rent_roll_date, "%Y-%m-%d"))
        except Exception:
            pass

        DATA_START = 9
        seen_units = set()   # Deduplicate: first occurrence of each unit_no wins
        write_idx  = 0       # Separate counter so duplicates don't consume a row slot
        for u in units:
            if write_idx >= (620 - DATA_START + 1):
                break
            unit_key = str(u.get("unit_no") or "").strip()
            if unit_key and unit_key in seen_units:
                continue    # Skip duplicate unit — future-tenant re-entry for same unit
            if unit_key:
                seen_units.add(unit_key)
            r = DATA_START + write_idx
            write_idx += 1

            # C — Unit No
            xml = _replace(xml, 3, r, u.get("unit_no"))

            # D — Unit Type
            xml = _replace(xml, 4, r, u.get("unit_type") or "")

            # H — Size (sqft)
            sqft = u.get("sqft")
            xml = _replace(xml, 8, r, int(sqft) if sqft else None)

            # I J K — Move-in, Lease Start, Lease End
            for col, field in [(9,"move_in"),(10,"lease_start"),(11,"lease_end")]:
                v = u.get(field)
                dt = None
                if v:
                    try: dt = datetime.strptime(v, "%m-%d-%Y")
                    except Exception: pass
                xml = _replace(xml, col, r, dt)

            # N — Resident Name (vacant/model/down units get a label, never blank)
            name   = u.get("resident_name") or ""
            status = str(u.get("status") or "").lower()
            if not name:
                if "model"   in status: name = "Model"
                elif "down"  in status or "offline" in status: name = "Down"
                else: name = "Vacant"
            xml = _replace(xml, 14, r, name)

            # O — Market Rent
            mkt = u.get("market_rent")
            xml = _replace(xml, 15, r, round(mkt) if mkt else None)

            # Q — Effective Rent (blank if zero/none)
            eff = u.get("effective_rent")
            xml = _replace(xml, 17, r, round(eff) if (eff and eff > 0) else None)

        tpl['xl/worksheets/sheet1.xml'] = xml.encode('utf-8')

        # Force calculation mode to automatic so formulas recalculate on open
        import re as _re2
        wb_xml = tpl['xl/workbook.xml'].decode('utf-8')
        wb_xml = wb_xml.replace('calcMode="manual"', 'calcMode="auto"')
        wb_xml = wb_xml.replace("calcMode='manual'", 'calcMode="auto"')

        # Strip the 2700+ junk definedNames inherited from old financial models.
        # They cause Excel to attempt broken cross-workbook lookups on open, which
        # triggers repair mode and prevents automatic recalculation.
        # Keep ONLY _xlnm._FilterDatabase (the autofilter range for the sheet).
        import re as _re3
        # Extract the one name we need
        _keep = _re3.findall(
            r'<definedName[^>]+_xlnm\._FilterDatabase[^>]*>[^<]*</definedName>',
            wb_xml
        )
        # Replace the entire <definedNames>...</definedNames> block
        _kept_block = '<definedNames>' + ''.join(_keep) + '</definedNames>'
        wb_xml = _re3.sub(
            r'<definedNames>(?:(?!<definedNames>)[\s\S])*?</definedNames>',
            _kept_block,
            wb_xml,
            count=1
        )
        tpl['xl/workbook.xml'] = wb_xml.encode('utf-8')

        # Remove stale calcChain.xml so Excel rebuilds it fresh on open.
        # A stale chain makes Excel skip recalculation even with calcMode="auto".
        tpl.pop('xl/calcChain.xml', None)
        if '[Content_Types].xml' in tpl:
            _ct = tpl['[Content_Types].xml'].decode('utf-8')
            _ct = _ct.replace(
                '<Override PartName="/xl/calcChain.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/>',
                ''
            )
            tpl['[Content_Types].xml'] = _ct.encode('utf-8')

        out = _io.BytesIO()
        with _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
            for name, data in tpl.items():
                zout.writestr(name, data)
        out.seek(0)
        return out.read()


    def run(self, raw_bytes, filename, rent_roll_date):
        self.log("step", f"🚀 Agent started — {filename}")
        try:
            raw_wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception as e:
            return {"ok":False,"error":f"Cannot open file: {e}"}

        sheet = self._find_raw_sheet(raw_wb)
        ws    = raw_wb[sheet]
        self.log("action", f"📄 Sheet: '{sheet}'  ({ws.max_row}r × {ws.max_column}c)")
        rows  = self._to_rows(ws)

        units       = []
        parser_name = None

        # Try known parsers
        for p in self.PARSERS:
            if p.can_handle(rows, filename):
                self.log("think", f"🔍 Matched parser: {p.name}")
                try:
                    units = p.extract(rows)
                    parser_name = p.name
                    self.log("action", f"   → {len(units)} units extracted")
                    break
                except Exception as e:
                    self.log("warn", f"   Parser failed: {e} — trying next")

        # AI fallback
        ai = AIFallbackParser()
        if not units:
            self.log("think", "🧠 No known format matched — invoking Claude for schema detection")
            schema = ai.detect_schema(rows, filename)
            self.log("action", f"   Format: {schema.get('format')} | {schema.get('structure')}")
            units       = ai.extract(rows, schema)
            parser_name = f"AI ({schema.get('format','?')})"
            self.log("action", f"   → {len(units)} units extracted")

        # Validate
        v = self._validate(units)
        self.log("ok" if v["ok"] else "warn",
                 ("✅ " if v["ok"] else "⚠️ ") + f"Validation: {v['msg']}")

        # Self-correction: if validation failed, try AI override
        if not v["ok"]:
            self.log("think", "🔄 Attempting AI self-correction...")
            try:
                schema   = ai.detect_schema(rows, filename)
                units2   = ai.extract(rows, schema)
                v2       = self._validate(units2)
                if v2["ok"] or len(units2) > len(units):
                    units, v = units2, v2
                    parser_name += " [corrected]"
                    self.log("ok", f"✅ Correction: {v['msg']}")
                else:
                    self.log("warn", "Correction did not improve — keeping original")
            except Exception as e:
                self.log("warn", f"Correction error: {e}")

        # Write template
        try:
            out_bytes = self._write_template(units, rent_roll_date)
            self.log("ok", "📥 Template populated successfully")
        except Exception as e:
            return {"ok":False,"error":f"Template write failed: {e}"}

        occ   = sum(1 for u in units if u.get("effective_rent"))
        t_eff = sum(u.get("effective_rent",0) or 0 for u in units)
        t_mkt = sum(u.get("market_rent",0) or 0 for u in units)

        return {
            "ok": True,
            "output_bytes": out_bytes,
            "metrics": {
                "total_units": len(units),
                "occupied": occ,
                "vacant": len(units) - occ,
                "total_effective_rent": t_eff,
                "total_market_rent": t_mkt,
                "parser": parser_name,
            },
            "validation": v,
            "preview": units[:5],
        }


# ─── Streamlit UI ─────────────────────────────────────────────────────────────
def main():
    st.title("🏢 Rent Roll AI Agent")
    st.caption("Auto-detects any rent roll format · extracts & transforms data · populates your standard template")
    st.divider()

    with st.sidebar:
        st.header("⚙️ Settings")
        rr_date = st.date_input("Rent Roll As-Of Date", value=datetime.today()).strftime("%Y-%m-%d")
        st.divider()
        st.markdown("**Built-in parsers**")
        for n in ["Yardi Voyager/Breeze","RealPage OneSite","MRI Living","AppFolio","Rent Manager"]:
            st.markdown(f"• {n}")
        st.markdown("• *Any other via Claude AI*")
        st.divider()
        st.markdown("**Agent logic**")
        st.markdown("""
- 🔍 Format auto-detection
- 💸 Rent code identification
- 🏗️ Building/section grouping
- 🚫 Skips applicants & future
- 🏚️ Zeros vacant/model/down
- ✅ Self-validates results
- 🔄 Claude correction loop
        """)

    uploaded = st.file_uploader("📂 Upload Raw Rent Roll (.xlsx)", type=["xlsx"])
    if not uploaded:
        st.info("Upload a raw rent roll file to begin.")
        return

    st.success(f"**{uploaded.name}** — {uploaded.size:,} bytes")
    _, mid, _ = st.columns([1,2,1])
    with mid:
        run_btn = st.button("🚀 Run Agent", type="primary", use_container_width=True)
    if not run_btn: return

    log_box = st.empty()
    log_lines = []

    def add_log(level, msg):
        cls  = {"step":"log-step","think":"log-think","action":"log-action",
                "warn":"log-warn","ok":"log-ok"}.get(level,"")
        icon = {"step":"▶","think":"💭","action":"⚡","warn":"⚠","ok":"✓"}.get(level,"·")
        ts   = datetime.now().strftime("%H:%M:%S")
        log_lines.append(f'<span class="{cls}">[{ts}] {icon} {msg}</span>')
        log_box.markdown(
            f'<div class="log-box">{"<br>".join(log_lines)}</div>',
            unsafe_allow_html=True,
        )

    raw_bytes = uploaded.read()
    with st.spinner(""):
        result = RentRollAgent(log_fn=add_log).run(raw_bytes, uploaded.name, rr_date)

    st.divider()
    if not result["ok"]:
        st.error(f"❌ {result.get('error','Unknown error')}")
        return

    m = result["metrics"]
    v = result["validation"]

    st.subheader("📊 Results")
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Total Units",   m["total_units"])
    c2.metric("Occupied",      m["occupied"])
    c3.metric("Vacant/Other",  m["vacant"])
    c4.metric("Total Eff Rent",f"${m['total_effective_rent']:,.0f}")
    c5.metric("Total Mkt Rent",f"${m['total_market_rent']:,.0f}")

    vi = "✅" if v["ok"] else "⚠️"
    st.info(f"{vi} **Validation:** {v['msg']}")
    st.caption(f"Parser: **{m['parser']}**")

    if result.get("preview"):
        with st.expander("🔎 Preview first 5 units"):
            df = pd.DataFrame(result["preview"])
            show = [c for c in ["unit_no","unit_type","sqft","resident_name",
                                 "move_in","lease_start","lease_end",
                                 "market_rent","effective_rent"] if c in df.columns]
            st.dataframe(df[show], use_container_width=True)

    st.divider()
    out_name = uploaded.name.replace(".xlsx","") + "_Rent_Roll.xlsx"
    st.download_button(
        "📥 Download Filled Rent Roll Template",
        data=result["output_bytes"],
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
